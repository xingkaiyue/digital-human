from __future__ import annotations

import csv
import uuid
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from .schema import POIRecord


class POIExtractor:
    """
    支持 xlsx / csv 的景点结构抽取
    约定字段可映射：
    - 景点名称 / name
    - 别名 / aliases
    - 景区名称 / scenic_name
    - 地址 / address
    - 简介 / description
    - 标签 / tags
    - 分类 / category
    """

    COLUMN_ALIASES = {
        "name": ["景点名称", "name", "poi_name"],
        "aliases": ["别名", "aliases"],
        "scenic_name": ["景区名称", "scenic_name"],
        "address": ["地址", "位置", "address"],
        "description": ["简介", "描述", "介绍", "description"],
        "tags": ["标签", "tags"],
        "category": ["分类", "category"],
        "scenic_id": ["景区ID", "scenic_id"],
    }

    def extract(self, file_path: str) -> List[POIRecord]:
        suffix = Path(file_path).suffix.lower()
        if suffix == ".csv":
            return self._extract_csv(file_path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._extract_xlsx(file_path)
        raise ValueError(f"暂不支持的文件类型: {suffix}")

    def _extract_csv(self, file_path: str) -> List[POIRecord]:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        return self._rows_to_records(rows, source_file=file_path)

    def _extract_xlsx(self, file_path: str) -> List[POIRecord]:
        wb = load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(x).strip() if x is not None else "" for x in rows[0]]
        data_rows = []
        for row in rows[1:]:
            item = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    item[headers[idx]] = "" if value is None else str(value).strip()
            data_rows.append(item)

        return self._rows_to_records(data_rows, source_file=file_path)

    def _rows_to_records(self, rows: List[dict], source_file: str) -> List[POIRecord]:
        result: List[POIRecord] = []

        for idx, row in enumerate(rows, start=2):
            name = self._pick(row, "name")
            if not name:
                continue

            aliases = self._split_multi(self._pick(row, "aliases"))
            tags = self._split_multi(self._pick(row, "tags"))

            scenic_name = self._pick(row, "scenic_name")
            scenic_id = self._pick(row, "scenic_id")
            category = self._pick(row, "category")
            address = self._pick(row, "address")
            description = self._pick(row, "description")

            poi_id = str(uuid.uuid5(
                uuid.NAMESPACE_DNS,
                f"{scenic_name}|{name}|{address}"
            ))

            result.append(
                POIRecord(
                    poi_id=poi_id,
                    scenic_id=scenic_id or None,
                    scenic_name=scenic_name or None,
                    name=name,
                    aliases=aliases,
                    category=category or None,
                    address=address or None,
                    description=description or None,
                    tags=tags,
                    source_file=Path(source_file).name,
                    source_row_no=idx,
                )
            )
        return result

    def _pick(self, row: dict, logical_name: str) -> str:
        for key in self.COLUMN_ALIASES.get(logical_name, []):
            value = row.get(key)
            if value:
                return str(value).strip()
        return ""

    @staticmethod
    def _split_multi(text: str) -> List[str]:
        if not text:
            return []
        raw = str(text).replace("，", ",").replace("；", ",").replace("/", ",")
        return [x.strip() for x in raw.split(",") if x.strip()]
